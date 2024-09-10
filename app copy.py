from flask import Flask, render_template, request, jsonify
import openpyxl
from openpyxl import Workbook
import os
import datetime
import json
from collections import defaultdict
import logging
from abc import ABC, abstractmethod
import calendar

app = Flask(__name__)

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Abstract base class for expenses
class BaseExpense(ABC):
    def __init__(self, category, date, amount):
        self.category = category
        self.date = datetime.datetime.strptime(date, "%Y-%m-%d").date()
        self.amount = float(amount)

    @abstractmethod
    def to_dict(self):
        pass

# Regular Expense class
class Expense(BaseExpense):
    def to_dict(self):
        return {
            "category": self.category,
            "date": self.date.strftime("%Y-%m-%d"),
            "amount": self.amount
        }

# Recurring Expense class that inherits from Expense
class RecurringExpense(Expense):
    def __init__(self, category, date, amount, frequency):
        super().__init__(category, date, amount)
        self.frequency = frequency

    def to_dict(self):
        expense_dict = super().to_dict()
        expense_dict["frequency"] = self.frequency
        return expense_dict

class ExpenseManager:
    def __init__(self, file_path):
        self.file_path = file_path
        self.expenses = []
        self.load_expenses()

    def load_expenses(self):
        try:
            if not os.path.exists(self.file_path):
                self.create_new_file()

            wb = openpyxl.load_workbook(self.file_path)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) >= 3:
                    self.expenses.append(Expense(row[0], row[1], row[2]))
        except Exception as e:
            logger.error(f"Error loading expenses: {str(e)}")

    def create_new_file(self):
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Category", "Date", "Amount"])
        wb.save(self.file_path)

    def save_expenses(self):
        try:
            wb = Workbook()
            sheet = wb.active
            sheet.append(["Category", "Date", "Amount"])

            for expense in self.expenses:
                sheet.append([expense.category, expense.date.strftime("%Y-%m-%d"), expense.amount])

            wb.save(self.file_path)
        except Exception as e:
            logger.error(f"Error saving expenses: {str(e)}")

    def add_expense(self, expense):
        self.expenses.append(expense)
        self.save_expenses()

    def update_expense(self, index, updated_expense):
        if 0 <= index < len(self.expenses):
            self.expenses[index] = updated_expense
            self.save_expenses()
        else:
            raise IndexError("Invalid expense index")

    def delete_expense(self, index):
        if 0 <= index < len(self.expenses):
            del self.expenses[index]
            self.save_expenses()
        else:
            raise IndexError("Invalid expense index")

    def get_expenses(self):
        return [expense.to_dict() for expense in self.expenses]

    def get_expenses_by_week(self):
        expenses_by_week = defaultdict(lambda: defaultdict(float))
        for expense in self.expenses:
            year, week, _ = expense.date.isocalendar()
            week_start = expense.date - datetime.timedelta(days=expense.date.weekday())
            week_key = f"{year}-W{week:02d} ({week_start.strftime('%m-%d')})"
            expenses_by_week[week_key][expense.category] += expense.amount
        return dict(expenses_by_week)

    def get_total_expenses(self):
        totals = defaultdict(float)
        for expense in self.expenses:
            totals[expense.category] += expense.amount
        return dict(totals)

    def clear_all_data(self):
        self.expenses = []
        self.save_expenses()

expense_manager = ExpenseManager("expenses.xlsx")

EXPENSE_CATEGORIES = [
    {"name": "Indoor Cooking", "key": "indoorCooking", "budget": {"good": 12500, "normal": 15000}},
    {"name": "Outdoor Dinners", "key": "outdoorDinners", "budget": {"good": 3250, "normal": 3750}},
    {"name": "Transport Fees", "key": "transportFees", "budget": {"good": 3500, "normal": 4000}},
    {"name": "Entertainment", "key": "entertainment", "budget": {"good": 3500, "normal": 4000}},
    {"name": "Education", "key": "education", "budget": {"good": 375, "normal": 500}},
    {"name": "Shopping", "key": "shopping", "budget": {"good": 5000, "normal": 6250}},
    {"name": "Medical Fees", "key": "medicalFees", "budget": {"good": 2000, "normal": 2500}},
]

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/expenses', methods=['GET', 'POST', 'DELETE'])
def handle_expenses():
    if request.method == 'GET':
        return jsonify(expense_manager.get_expenses())
    elif request.method == 'POST':
        try:
            new_expense_data = request.json
            new_expense = Expense(**new_expense_data)
            expense_manager.add_expense(new_expense)
            return jsonify({"status": "success"})
        except Exception as e:
            logger.error(f"Error adding expense: {str(e)}")
            return jsonify({"status": "error", "message": str(e)}), 400
    elif request.method == 'DELETE':
        try:
            expense_manager.clear_all_data()
            return jsonify({"status": "success"})
        except Exception as e:
            logger.error(f"Error clearing all data: {str(e)}")
            return jsonify({"status": "error", "message": str(e)}), 400

@app.route('/api/expenses/<int:index>', methods=['PUT', 'DELETE'])
def handle_expense(index):
    try:
        if request.method == 'PUT':
            updated_expense_data = request.json
            updated_expense = Expense(**updated_expense_data)
            expense_manager.update_expense(index, updated_expense)
            return jsonify({"status": "success"})
        elif request.method == 'DELETE':
            expense_manager.delete_expense(index)
            return jsonify({"status": "success"})
    except IndexError:
        return jsonify({"status": "error", "message": "Invalid expense index"}), 404
    except Exception as e:
        logger.error(f"Error handling expense: {str(e)}")
        return jsonify({"status": "error", "message": str(e)}), 400

@app.route('/api/budget_status', methods=['GET'])
def get_budget_status():
    weekly_expenses = expense_manager.get_expenses_by_week()
    total_expenses = expense_manager.get_total_expenses()

    budget_status = {
        "weekly": weekly_expenses,
        "total": {
            category['key']: {
                "amount": total_expenses.get(category['key'], 0),
                "status": get_status(total_expenses.get(category['key'], 0), category['budget'])
            }
            for category in EXPENSE_CATEGORIES
        }
    }

    return jsonify(budget_status)

def get_status(amount, budget):
    if amount < budget['good']:
        return "Good"
    elif amount < budget['normal']:
        return "Normal"
    else:
        return "Over Budget"

if __name__ == '__main__':
    app.run(debug=True)
