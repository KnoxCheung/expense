import os
import json
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.chart import PieChart, Reference
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

class FinancialEntity:
    def __init__(self, amount):
        self.amount = amount
        self.date = datetime.now().strftime("%Y-%m-%d")

    def to_dict(self):
        return {
            "amount": self.amount,
            "date": self.date
        }

class Expense(FinancialEntity):
    def __init__(self, category, amount):
        super().__init__(amount)
        self.category = category

    def to_dict(self):
        return {
            **super().to_dict(),
            "category": self.category
        }

class RecurringExpense(Expense):
    def __init__(self, category, amount, frequency):
        super().__init__(category, amount)
        self.frequency = frequency  # 'daily', 'weekly', 'monthly', 'yearly'

    def to_dict(self):
        return {
            **super().to_dict(),
            "frequency": self.frequency
        }

class Budget(FinancialEntity):
    def __init__(self, limit):
        super().__init__(limit)
        self.spent = 0

    def add_expense(self, amount):
        self.spent += amount

    def get_remaining(self):
        return self.amount - self.spent

    def to_dict(self):
        return {
            **super().to_dict(),
            "spent": self.spent,
            "remaining": self.get_remaining()
        }

class ExpenseManager:
    def __init__(self, filename):
        self.filename = filename
        self.expenses = []
        self.recurring_expenses = []
        self.budget = Budget(1000)  # Default budget limit
        self.load_data()

    def add_expense(self, category, amount):
        expense = Expense(category, amount)
        self.expenses.append(expense)
        self.budget.add_expense(amount)
        self.save_data()

    def add_recurring_expense(self, category, amount, frequency):
        recurring_expense = RecurringExpense(category, amount, frequency)
        self.recurring_expenses.append(recurring_expense)
        self.save_data()

    def get_expenses(self):
        return self.expenses

    def get_recurring_expenses(self):
        return self.recurring_expenses

    def clear_data(self):
        self.expenses = []
        self.recurring_expenses = []
        self.budget = Budget(1000)
        self.save_data()

    def get_expense_summary(self):
        summary = {}
        for expense in self.expenses:
            if expense.category in summary:
                summary[expense.category] += expense.amount
            else:
                summary[expense.category] = expense.amount
        return summary

    def get_budget_status(self):
        return self.budget.to_dict()

    def load_data(self):
        try:
            if not os.path.exists(self.filename):
                self.save_data()  # Create a new file if it doesn't exist
                return

            wb = load_workbook(self.filename)
            
            expenses_sheet = wb['Expenses']
            self.expenses = []
            for row in expenses_sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    self.expenses.append(Expense(row[0], row[1]))
                    self.budget.add_expense(row[1])

            recurring_sheet = wb['RecurringExpenses']
            self.recurring_expenses = []
            for row in recurring_sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    self.recurring_expenses.append(RecurringExpense(row[0], row[1], row[2]))

            budget_sheet = wb['Budget']
            self.budget = Budget(budget_sheet['A1'].value)
            self.budget.spent = budget_sheet['B1'].value

        except Exception as e:
            print(f"Error loading data: {e}")
            self.expenses = []
            self.recurring_expenses = []
            self.budget = Budget(1000)

    def save_data(self):
        try:
            wb = Workbook()
            
            expenses_sheet = wb.active
            expenses_sheet.title = "Expenses"
            expenses_sheet.append(["Category", "Amount", "Date"])
            for expense in self.expenses:
                expenses_sheet.append([expense.category, expense.amount, expense.date])

            recurring_sheet = wb.create_sheet(title="RecurringExpenses")
            recurring_sheet.append(["Category", "Amount", "Frequency", "Date"])
            for expense in self.recurring_expenses:
                recurring_sheet.append([expense.category, expense.amount, expense.frequency, expense.date])

            budget_sheet = wb.create_sheet(title="Budget")
            budget_sheet['A1'] = self.budget.amount
            budget_sheet['B1'] = self.budget.spent

            wb.save(self.filename)
        except Exception as e:
            print(f"Error saving data: {e}")

    def generate_pie_chart(self):
        summary = self.get_expense_summary()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Expense Summary"

        ws.append(["Category", "Amount"])
        for category, amount in summary.items():
            ws.append([category, amount])

        pie = PieChart()
        labels = Reference(ws, min_col=1, min_row=2, max_row=len(summary) + 1)
        data = Reference(ws, min_col=2, min_row=1, max_row=len(summary) + 1)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Expense Distribution"

        ws.add_chart(pie, "E1")

        chart_filename = "expense_chart.xlsx"
        wb.save(chart_filename)
        return chart_filename

    def send_weekly_reminder(self, email):
        try:
            summary = self.get_expense_summary()
            budget_status = self.get_budget_status()

            message = MIMEMultipart()
            message['Subject'] = 'Weekly Expense Reminder'
            message['From'] = 'your_email@example.com'
            message['To'] = email

            body = f"""
            Weekly Expense Summary:
            {json.dumps(summary, indent=2)}

            Budget Status:
            Limit: ${budget_status['amount']}
            Spent: ${budget_status['spent']}
            Remaining: ${budget_status['remaining']}

            Recurring Expenses:
            {json.dumps([e.to_dict() for e in self.recurring_expenses], indent=2)}
            """

            message.attach(MIMEText(body, 'plain'))

            with smtplib.SMTP('smtp.gmail.com', 587) as server:
                server.starttls()
                server.login('your_email@example.com', 'your_password')
                server.send_message(message)

            print("Weekly reminder sent successfully")
        except Exception as e:
            print(f"Error sending weekly reminder: {e}")

# Usage example
if __name__ == "__main__":
    manager = ExpenseManager("expenses.xlsx")

    # Add some sample expenses
    manager.add_expense("Groceries", 50.0)
    manager.add_expense("Transport", 30.0)
    manager.add_expense("Entertainment", 100.0)

    # Add a recurring expense
    manager.add_recurring_expense("Rent", 1000.0, "monthly")

    # Print expense summary
    print("Expense Summary:")
    print(json.dumps(manager.get_expense_summary(), indent=2))

    # Print budget status
    print("\nBudget Status:")
    print(json.dumps(manager.get_budget_status(), indent=2))

    # Print recurring expenses
    print("\nRecurring Expenses:")
    print(json.dumps([e.to_dict() for e in manager.get_recurring_expenses()], indent=2))

    # Generate pie chart
    chart_file = manager.generate_pie_chart()
    print(f"\nPie chart generated: {chart_file}")

    # Send weekly reminder
    manager.send_weekly_reminder("recipient@example.com")