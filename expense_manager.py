import os
import json
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.chart import PieChart, Reference
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import calendar

class FinancialEntity:
    def __init__(self, amount):
        self.amount = amount
        self.date = datetime.now().strftime("%Y-%m-%d")

    def to_dict(self):
        return {
            "amount": self.amount,
            "date": self.date
        }

class Expense(FinancialEntity):
    def __init__(self, category, amount, date):
        super().__init__(amount)
        self.category = category
        self.date = date

    def to_dict(self):
        return {
            **super().to_dict(),
            "category": self.category,
            "date": self.date
        }

class RecurringExpense(Expense):
    def __init__(self, category, amount, frequency, date):
        super().__init__(category, amount, date)
        self.frequency = frequency  # 'daily', 'weekly', 'monthly', 'yearly'

    def to_dict(self):
        return {
            **super().to_dict(),
            "frequency": self.frequency
        }

class Budget:
    def __init__(self):
        self.limits = {
            "indoorCooking": {"good": 12500, "normal": 15000},
            "outdoorDinners": {"good": 3250, "normal": 3750},
            "transportFees": {"good": 3500, "normal": 4000},
            "entertainment": {"good": 3500, "normal": 4000},
            "education": {"good": 375, "normal": 500},
            "shopping": {"good": 5000, "normal": 6250},
            "medicalFees": {"good": 2000, "normal": 2500}
        }

    def update_limit(self, category, good_limit, normal_limit):
        if category in self.limits:
            self.limits[category] = {"good": good_limit, "normal": normal_limit}

    def get_limit(self, category):
        return self.limits.get(category, {"good": 0, "normal": 0})

    def to_dict(self):
        return self.limits

class ExpenseManager:
    def __init__(self, filename):
        self.filename = filename
        self.expenses = []
        self.recurring_expenses = []
        self.budget = Budget()
        self.load_data()

    def add_expense(self, category, amount, date):
        try:
            expense = Expense(category, float(amount), date)
            self.expenses.append(expense)
            self.save_data()
            return True
        except ValueError:
            print("Invalid amount. Please enter a valid number.")
            return False

    def add_recurring_expense(self, category, amount, frequency, date):
        try:
            recurring_expense = RecurringExpense(category, float(amount), frequency, date)
            self.recurring_expenses.append(recurring_expense)
            self.save_data()
            return True
        except ValueError:
            print("Invalid amount. Please enter a valid number.")
            return False

    def remove_expense(self, index):
        if 0 <= index < len(self.expenses):
            del self.expenses[index]
            self.save_data()
            return True
        return False

    def remove_recurring_expense(self, index):
        if 0 <= index < len(self.recurring_expenses):
            del self.recurring_expenses[index]
            self.save_data()
            return True
        return False

    def edit_expense(self, index, category, amount, date):
        if 0 <= index < len(self.expenses):
            try:
                self.expenses[index] = Expense(category, float(amount), date)
                self.save_data()
                return True
            except ValueError:
                print("Invalid amount. Please enter a valid number.")
                return False
        return False

    def edit_recurring_expense(self, index, category, amount, frequency, date):
        if 0 <= index < len(self.recurring_expenses):
            try:
                self.recurring_expenses[index] = RecurringExpense(category, float(amount), frequency, date)
                self.save_data()
                return True
            except ValueError:
                print("Invalid amount. Please enter a valid number.")
                return False
        return False

    def get_expenses(self):
        return self.expenses

    def get_recurring_expenses(self):
        return self.recurring_expenses

    def clear_data(self):
        self.expenses = []
        self.recurring_expenses = []
        self.save_data()

    def get_expense_summary(self):
        summary = {}
        for expense in self.expenses:
            if expense.category in summary:
                summary[expense.category] += expense.amount
            else:
                summary[expense.category] = expense.amount
        return summary

    def get_budget_status(self):
        return self.budget.to_dict()

    def update_budget_limit(self, category, good_limit, normal_limit):
        self.budget.update_limit(category, float(good_limit), float(normal_limit))
        self.save_data()

    def load_data(self):
        try:
            if not os.path.exists(self.filename):
                self.save_data()  # Create a new file if it doesn't exist
                return

            wb = load_workbook(self.filename)

            expenses_sheet = wb['Expenses']
            self.expenses = []
            for row in expenses_sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    self.expenses.append(Expense(row[0], row[1], row[2]))

            recurring_sheet = wb['RecurringExpenses']
            self.recurring_expenses = []
            for row in recurring_sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    self.recurring_expenses.append(RecurringExpense(row[0], row[1], row[2], row[3]))

            budget_sheet = wb['Budget']
            for row in budget_sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    self.budget.update_limit(row[0], row[1], row[2])

        except Exception as e:
            print(f"Error loading data: {e}")
            self.expenses = []
            self.recurring_expenses = []

    def save_data(self):
        try:
            wb = Workbook()

            expenses_sheet = wb.active
            expenses_sheet.title = "Expenses"
            expenses_sheet.append(["Category", "Amount", "Date"])
            for expense in self.expenses:
                expenses_sheet.append([expense.category, expense.amount, expense.date])

            recurring_sheet = wb.create_sheet(title="RecurringExpenses")
            recurring_sheet.append(["Category", "Amount", "Frequency", "Date"])
            for expense in self.recurring_expenses:
                recurring_sheet.append([expense.category, expense.amount, expense.frequency, expense.date])

            budget_sheet = wb.create_sheet(title="Budget")
            budget_sheet.append(["Category", "Good Limit", "Normal Limit"])
            for category, limits in self.budget.to_dict().items():
                budget_sheet.append([category, limits["good"], limits["normal"]])

            wb.save(self.filename)
        except Exception as e:
            print(f"Error saving data: {e}")

    def generate_pie_chart(self):
        summary = self.get_expense_summary()

        wb = Workbook()
        ws = wb.active
        ws.title = "Expense Summary"

        ws.append(["Category", "Amount"])
        for category, amount in summary.items():
            ws.append([category, amount])

        pie = PieChart()
        labels = Reference(ws, min_col=1, min_row=2, max_row=len(summary) + 1)
        data = Reference(ws, min_col=2, min_row=1, max_row=len(summary) + 1)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Expense Distribution"

        ws.add_chart(pie, "E1")

        chart_filename = "expense_chart.xlsx"
        wb.save(chart_filename)
        return chart_filename

    def send_weekly_reminder(self, email):
        try:
            summary = self.get_expense_summary()
            budget_status = self.get_budget_status()

            message = MIMEMultipart()
            message['Subject'] = 'Weekly Expense Reminder'
            message['From'] = 'your_email@example.com'
            message['To'] = email

            body = f"""
            Weekly Expense Summary:
            {json.dumps(summary, indent=2)}

            Budget Status:
            {json.dumps(budget_status, indent=2)}

            Recurring Expenses:
            {json.dumps([e.to_dict() for e in self.recurring_expenses], indent=2)}
            """

            message.attach(MIMEText(body, 'plain'))

            with smtplib.SMTP('smtp.gmail.com', 587) as server:
                server.starttls()
                server.login('your_email@example.com', 'your_password')
                server.send_message(message)

            print("Weekly reminder sent successfully")
        except Exception as e:
            print(f"Error sending weekly reminder: {e}")

    def get_expenses_for_range(self, start_date, end_date):
        expenses_in_range = []
        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date, "%Y-%m-%d")

        for expense in self.expenses:
            expense_date = datetime.strptime(expense.date, "%Y-%m-%d")
            if start <= expense_date <= end:
                expenses_in_range.append(expense)

        return expenses_in_range

    def get_weekly_expense_alarm(self):
        today = datetime.now()
        start_of_week = today - timedelta(days=today.weekday())
        end_of_week = start_of_week + timedelta(days=6)

        weekly_expenses = self.get_expenses_for_range(start_of_week.strftime("%Y-%m-%d"), end_of_week.strftime("%Y-%m-%d"))

        category_totals = {}
        for expense in weekly_expenses:
            if expense.category in category_totals:
                category_totals[expense.category] += expense.amount
            else:
                category_totals[expense.category] = expense.amount

        alarms = []
        for category, total in category_totals.items():
            limit = self.budget.get_limit(category)
            if total > limit["normal"]:
                alarms.append(f"Warning: {category} expenses (${total:.2f}) have exceeded the normal limit (${limit['normal']:.2f})")
            elif total > limit["good"]:
                alarms.append(f"Caution: {category} expenses (${total:.2f}) have exceeded the good limit (${limit['good']:.2f})")

        return alarms

    def get_monthly_calendar(self, year, month):
        month_expenses = {}
        _, last_day = calendar.monthrange(year, month)
        for day in range(1, last_day + 1):
            date_str = f"{year}-{month:02d}-{day:02d}"
            day_expenses = [e for e in self.expenses if e.date == date_str]
            if day_expenses:
                month_expenses[day] = day_expenses
        return month_expenses

# Usage example
if __name__ == "__main__":
    manager = ExpenseManager("expenses.xlsx")

    # Add some sample expenses
    manager.add_expense("indoorCooking", 50.0, "2023-05-01")
    manager.add_expense("transportFees", 30.0, "2023-05-02")
    manager.add_expense("entertainment", 100.0, "2023-05-03")

    # Add a recurring expense
    manager.add_recurring_expense("medicalFees", 1000.0, "monthly", "2023-05-01")

    # Print expense summary
    print("Expense Summary:")
    print(json.dumps(manager.get_expense_summary(), indent=2))

    # Print budget status
    print("\nBudget Status:")
    print(json.dumps(manager.get_budget_status(), indent=2))

    # Print recurring expenses
    print("\nRecurring Expenses:")
    print(json.dumps([e.to_dict() for e in manager.get_recurring_expenses()], indent=2))

    # Generate pie chart
    chart_file = manager.generate_pie_chart()
    print(f"\nPie chart generated: {chart_file}")

    # Send weekly reminder
    manager.send_weekly_reminder("recipient@example.com")

    # Demonstrate use of range() function
    print("\nExpenses in the last 7 days:")
    today = datetime.now()
    seven_days_ago = today - timedelta(days=7)
    recent_expenses = manager.get_expenses_for_range(seven_days_ago.strftime("%Y-%m-%d"), today.strftime("%Y-%m-%d"))
    for expense in recent_expenses:
        print(f"{expense.date}: {expense.category} - ${expense.amount}")

    # Get weekly expense alarm
    print("\nWeekly Expense Alarms:")
    alarms = manager.get_weekly_expense_alarm()
    for alarm in alarms:
        print(alarm)

    # Get monthly calendar
    print("\nMonthly Calendar (May 2023):")
    month_expenses = manager.get_monthly_calendar(2023, 5)
    cal = calendar.monthcalendar(2023, 5)
    for week in cal:
        for day in week:
            if day == 0:
                print("   ", end=" ")
            else:
                expenses = month_expenses.get(day, [])
                total = sum(e.amount for e in expenses)
                print(f"{day:2d}${total:6.2f}", end=" ")
        print()
